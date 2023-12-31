
# -*- coding: latin-1 -*-
import os
import pandas as pd
import sys
import recursos 
import pdfplumber
import PyPDF2
import aspose.pdf as ap
import aspose.pdf as pdf
import openpyxl
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
from openpyxl.styles import PatternFill
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QPushButton, QFileDialog, QMainWindow, QVBoxLayout, QMessageBox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle
from ast import Lambda
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow, QGraphicsDropShadowEffect
from PyQt5.QtCore import QPropertyAnimation,QEasingCurve
from PyQt5.QtGui import QColor
from PyQt5.QtCore import QThread, pyqtSignal, QTimer, QEventLoop
from PyQt5.QtGui import QCloseEvent
from PyQt5.uic import loadUi
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtCore import QFile, QIODevice
from PyQt5 import uic

class rooti(QMainWindow):
    def __init__(self):
        super(rooti,self).__init__()
        script_dir = os.path.dirname(os.path.realpath(__file__))
        ui_file_path = os.path.join(script_dir, 'interfaz.ui')
        loadUi(ui_file_path,self)
        
        #condiciones iniciales botones
        self.metodo1.model().item(0).setEnabled(False)
        self.proceso1.model().item(0).setEnabled(False)
        self.proceso2.model().item(0).setEnabled(False)
        self.metodo1.setEnabled(False)
        # self.proceso1.setEnabled(False)
        self.proceso2.setEnabled(False)
        self.bt_filter1.setEnabled(False)
        self.bt_save1.setEnabled(False)
        self.bt_load1.setEnabled(False)
        self.bt_load2.setEnabled(False)
        self.bt_save2.setEnabled(False)
        self.bt_filter2.setEnabled(False)
        #botones de funciones
        self.bt_equi1.clicked.connect(self.equipo1)
        self.bt_equi2.clicked.connect(self.equipo2)
        self.bt_load1.clicked.connect(self.load1)
        self.bt_load2.clicked.connect(self.load2)
        self.bt_filter1.clicked.connect(self.filter1)
        self.bt_filter2.clicked.connect(self.filter2)
        self.bt_save1.clicked.connect(self.save1)
        self.bt_save2.clicked.connect(self.save2)
        #botones de barra de titulo
        self.bt_maxi.clicked.connect(self.maximizar)
        self.bt_restart.clicked.connect(self.normal)
        self.bt_mini.clicked.connect(self.minimizar)
        self.bt_close.clicked.connect(lambda: self.close())
        
        #eliminar la ventana del main
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setWindowOpacity(1)
        #SizeGrip
        self.gripSize=10
        self.grip=QtWidgets.QSizeGrip(self)
        self.grip.resize(self.gripSize,self.gripSize)
        #mover ventana
        self.frame_top.mouseMoveEvent=self.mover_ventana
        #conexion botones
        self.bt_equi1.clicked.connect(lambda:self.stackedWidget.setCurrentWidget(self.page_equipo1))
        self.bt_equi2.clicked.connect(lambda:self.stackedWidget.setCurrentWidget(self.page_equipo2))
        self.metodo1.currentIndexChanged.connect(self.actualizarComboBoxProceso)
        self.proceso2.currentIndexChanged.connect(self.actualizarnewproces)

    @pyqtSlot()
    def actualizarComboBoxProceso(self):
            metodo_seleccionado = self.metodo1.currentText()
              # Limpiar las opciones actuales en comboBoxProceso1
            self.metodo=metodo_seleccionado
            if metodo_seleccionado == "Requerimientos Usales":
                
                self.etiqueta.show()
                self.newcolmns.hide()
                self.agregar.hide()
                self.label_7.setText("Has selecionado "+self.metodo)
                
            elif metodo_seleccionado == "Agregar columnas":
                self.newcolmns.show()
                self.agregar.show()
                self.proceso1.setEnabled(True)
                
                self.label_7.setText("Has selecionado "+self.metodo)
                
            
    
    @pyqtSlot()
    def actualizarnewproces(self):
            metodo_seleccionado = self.proceso2.currentText()
            self.newproces.hide()
            self.bt_ingresar.setEnabled(True)
            if metodo_seleccionado == "Otro":
                self.newproces.show()
                
                self.text=self.bt_ingresar.clicked.connect(self.copiarTexto)
                
               
    def copiarTexto(self):
        
        dato=self.dato1.text()
        self.dato2='Media '+dato+'  (mg/L)'
               
        
        print(self.dato2)
        
        
        
        
        
        
        

    

    def minimizar(self):
        self.showMinimized()
        
    def normal(self):
        self.showNormal()
        self.bt_restart.hide()
        self.bt_maxi.show()
    
    def maximizar(self): 
        self.showMaximized()
        self.bt_restart.show()
        self.bt_maxi.hide()
        
    def mousePressEvent(self,event):
        self.click_position=event.globalPos()
    
    def mover_ventana(self,event):
        if self.isMaximized()==False:
            if event.buttons()==QtCore.Qt.LeftButton:
                self.move(self.pos()+event.globalPos()-self.click_position)
                self.click_position=event.globalPos()
                event.accept()
        if event.globalPos().y()<=10:
            self.showMaximized()
            self.bt_maxi.hide()
            self.bt_restart.show()
        else:
            self.showNormal()
            self.bt_maxi.show()
            self.bt_restart.hide()
                           

 
    def mover_menu(self):
        if True:
            width=self.frame_control.width()
            normal=0
            if width==0:
                extender=300
            else:
                extender=normal
            self.animacion=QPropertyAnimation(self.frame_control,b'minimumWidth')
            self.animacion.setDuration(300)
            self.animacion.setStartValue(width)
            self.animacion.setEndValue(extender)
            self.animacion.setEasingCurve(QtCore.QEasingCurve.InOutQuart)
            self.animacion.start()
    
    def equipo1(self):
        self.bt_load1.setEnabled(True)
        
        # Conectar la se�al currentIndexChanged de comboBoxMetodo1 a la funci�n actualizarComboBoxProceso

        
    def load1(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        archivo, _ = QFileDialog.getOpenFileName(self, 'Buscador de csv', '', 'Archivos CSV (*.csv);;Todos los archivos (*)', options=options)
        if archivo:
            self.label_4.setText("Archivo cargado")
            self.archivo_seleccionado = archivo
            self.bt_filter1.setEnabled(True)
            self.metodo1.setEnabled(True)
            
                    
        else:
            self.label_4.setText('No se seleccion� ning�n archivo.')
            self.bt_filter1.setEnabled(False)
    
    def filter1(self):
        if self.archivo_seleccionado:
            try: 
                
                metho1=self.metodo1.currentText()
                process1=self.proceso1.currentText()
                df = pd.read_csv(self.archivo_seleccionado, encoding='latin-1')
                columnas = df.columns
                print(columnas)
                encontradas = []
                def tablita(df, variables):
                    encontradas = []
                    for var_buscada in variables:
                        encontrada = False
                        for columna in df.columns:
                            if columna == var_buscada:
                                encontradas.append(var_buscada)
                                encontrada = True
                                break
                        if not encontrada:
                            print(f"No encontrada: {var_buscada}")

                    if not encontradas:
                        return "Excel vac�o"  # Retorna un mensaje si no se encontraron variables

                    tabla = df[encontradas]
                    return tabla
                
                if metho1=='Requerimientos Usales':
                    variables=['Sample ID','Abs (Corr)','RSD (Corr Abs)','SD (Corr Abs)','Conc (Samp)','RSD (Conc)','Abs (Corr)1','Conc (Samp)1','Abs (Corr)2','Conc (Samp)2','Abs (Corr)3','Conc (Samp)3']
                    
                    self.tabla = tablita(df,variables)

                elif metho1=='Agregar columnas':
                    if process1=='SD':
                        variables=['Sample ID','Abs (Corr)','RSD (Corr Abs)','SD (Corr Abs)','Conc (Samp)','RSD (Conc)','Abs (Corr)1','Conc (Samp)1','Abs (Corr)2','Conc (Samp)2','Abs (Corr)3','Conc (Samp)3','SD (Calib)','SD (Samp)']
                        self.tabla = tablita(df,variables)

                    elif process1=='Sig':
                        variables=['Sample ID','Abs (Corr)','RSD (Corr Abs)','SD (Corr Abs)','Conc (Samp)','RSD (Conc)','Abs (Corr)1','Conc (Samp)1','Abs (Corr)2','Conc (Samp)2','Abs (Corr)3','Conc (Samp)3','Sig Area','Sig Ht']
                        self.tabla = tablita(df,variables)

                    elif process1=='Bkgnd':
                        variables=['Sample ID','Abs (Corr)','RSD (Corr Abs)','SD (Corr Abs)','Conc (Samp)','RSD (Conc)','Abs (Corr)1','Conc (Samp)1','Abs (Corr)2','Conc (Samp)2','Abs (Corr)3','Conc (Samp)3','Bkgnd Area','Bkgnd Ht','BkgndArea1','BkgndHt1','BkgndArea2','BkgndHt2','BkgndArea3','BkgndHt3']
                        self.tabla = tablita(df,variables)


                    elif process1=='Peak':
                       variables=['Sample ID','Abs (Corr)','RSD (Corr Abs)','SD (Corr Abs)','Conc (Samp)','RSD (Conc)','Abs (Corr)1','Conc (Samp)1','Abs (Corr)2','Conc (Samp)2','Abs (Corr)3','Conc (Samp)3','PeakArea1','PeakHt1','PeakArea2','PeakHt2','PeakArea3','PeakHt3']

                       self.tabla = tablita(df,variables)


                
                    
                self.label_4.setText("Archivo Filtrado")
                self.bt_save1.setEnabled(True) 
                    
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Error al filtrar la tabla: {str(e)}')
        else:
            QMessageBox.warning(self, 'Advertencia', 'No se seleccion� ning�n archivo.')
    
    
    def save1(self):
        if hasattr(self, 'tabla'):
            try:
                nombre_archivo = os.path.splitext(os.path.basename(self.archivo_seleccionado))[0]
                nuevo_nombre_excel = f"Filtrado_{nombre_archivo}.xlsx"
                options = QFileDialog.Options()
                options |= QFileDialog.ReadOnly
                ruta_guardado, _ = QFileDialog.getSaveFileName(self, 'Guardar en Excel', nuevo_nombre_excel, 'Archivos Excel (*.xls);;Todos los archivos (*)', options=options)
                
                if ruta_guardado:
                    df = pd.DataFrame(self.tabla)
                    writer = pd.ExcelWriter(ruta_guardado, engine='xlsxwriter')
                    df.to_excel(writer, index=False, sheet_name='Hoja1')
                    workbook  = writer.book
                    worksheet = writer.sheets['Hoja1']
                    header_format = workbook.add_format({'bg_color': '#00FF00', 'align': 'center', 'valign': 'vcenter','bold': True,'border':1})
                    for col_num, value in enumerate(df.columns.values):
                        worksheet.write(0, col_num, value, header_format)
                    cell_format=workbook.add_format({ 'align': 'center', 'valign': 'vcenter','border':1})
                    
                    for i, col in enumerate(df.columns):
                        max_len = max(df[col].astype(str).str.len())
                        max_len = max(max_len, len(col))
                        worksheet.set_column(i, i, max_len + 2)
                    worksheet.set_default_row(15)  # Altura de fila predeterminada
                    worksheet.conditional_format(1, 0, len(df), len(df.columns)-1, {'type': 'no_blanks', 'format': cell_format})
                    writer.close()
                    QMessageBox.information(self, 'Informaci�n', f'Archivo Excel guardado en: {ruta_guardado}')
                    self.label_4.setText("Archivo Guardado")
                    self.bt_save1.setEnabled(False)
                    self.bt_filter1.setEnabled(False)
                    self.label_7.setText('')
                    self.metodo1.setCurrentIndex(0)
                    self.proceso1.setCurrentIndex(0)
                   
                    
                else:
                    QMessageBox.warning(self, 'Advertencia', 'No se seleccion� una ruta de guardado.')
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Error al guardar en Excel: {str(e)}')
        else:
            QMessageBox.warning(self, 'Advertencia', 'No se seleccion� ninguna tabla.')
            
   

    def equipo2(self):
        self.bt_load2.setEnabled(True)
        self.newproces.hide()

    def load2(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(self, "Select PDF File", "", "PDF files (*.pdf)")
        self.archivo_seleccionado = file_path
        if file_path:
            self.label_6.setText("Archivo cargado")
            dataframes = []
            def extract_columns_from_table(table):
                num_columns = len(table[0])
                for col_idx in range(num_columns):
                    column_name = table[0][col_idx]
                    column = [row[col_idx] for row in table[1:]]
                    yield pd.Series(column,name=column_name)
            all_series = []
            with pdfplumber.open(file_path) as pdf:
                 for page in pdf.pages:
       
                    tables = page.extract_tables()
                    for table in tables:
                        for series in extract_columns_from_table(table):
                            all_series.append(series)
                            print(series)
                            print() 
                        
                        print("---- Fin de la tabla ----")
            df = pd.concat(all_series, axis=1)
            print(df)
            self.df = df
            columnas = df.columns
            self.bt_filter2.setEnabled(True)
            print(columnas)
            self.proceso2.setEnabled(True)
        else:
            self.label_6.setText("No se seleccion� ning�n archivo.")
            self.bt_filter2.setEnabled(False)
            self.proceso2.setEnabled(False)
            
            
            
                      

        
            
    def filter2(self):
        if hasattr(self, 'df'):
            self.bt_save2.setEnabled(True)
            
            try:
                
                 
                 
                 df = self.df
                 def tablita(df, variables):
                    encontradas = []
                    for var_buscada in variables:
                        encontrada = False
                        for columna in df.columns:
                            if columna == var_buscada:
                                encontradas.append(var_buscada)
                                encontrada = True
                                break
                        if not encontrada:
                            print(f"No encontrada: {var_buscada}")

                    if not encontradas:
                        return "Excel vac�o"  # Retorna un mensaje si no se encontraron variables

                    tabla = df[encontradas]
                    return tabla
                 proces3=self.proceso2.currentText()
                 columnas_seleccionadas = ['ID de muestra','Media Fosforo Total (mg/L)', 'Desv est', 'Media Abs 690 (AU)', 'Abs 690 Desv est',
                     'Triplicados 1', 'Triplicados 2', 'Triplicados 3',
                     'Abs 690 Triplicados 1', 'Abs 690 Triplicados 2', 'Abs 690 Triplicados 3']
                 if proces3=='Concentraci�n de F�sforo':
                     df[['Triplicados 1', 'Triplicados 2', 'Triplicados 3']] = df['Triplicados'].str.split(expand=True)
                     vari=['Abs 690 Triplicados']
                     self.tabla=tablita(df,vari)
                     print(self.tabla)
                     if self.df.empty:
                         self.tabla = tablita(df,columnas_seleccionadas)
                     else:
                         df[['Abs 690 Triplicados 1', 'Abs 690 Triplicados 2', 'Abs 690 Triplicados 3']] = df['Abs 690 Triplicados'].str.split(expand=True)
                         self.tabla = tablita(df,columnas_seleccionadas)
                     
                     
                 
                 elif proces3=='Clorofila':
                      
                      df[['Replicates 665nm', 'Replicates 665nm 2']] = df['Replicates 665nm'].str.split(expand=True)
                      print(df[['Replicates 665nm', 'Replicates 665nm 2']])
                      df[['Replicates 750nm 1', 'Replicates 750nm 2']] = df['Replicates 750nm'].str.split(expand=True)
                      print(df[['Replicates 750nm 1', 'Replicates 750nm 2']])
                      vari=['Replicates 750nm']
                      self.tabla=tablita(df,vari)
                      print(self.tabla)
                      if self.df.empty:
                         self.tabla = tablita(df,columnas)
                      else:
                        columnas = ['ID de muestra','Replicates 665nm', 'Replicates 665nm 2','Media de absorbancias665nm','Replicates 750nm 1', 'Replicates 750nm 2','Media de absorbancias750nm']
                        self.tabla = tablita(df,columnas)
                 elif proces3=='Otro':
                     
                     
                     df[['Triplicados 1', 'Triplicados 2', 'Triplicados 3']] = df['Triplicados'].str.split(expand=True)
                     vari=['Abs 690 Triplicados']
                     self.tabla=tablita(df,vari)
                     print(self.tabla)
                     if self.df.empty:
                         self.tabla = tablita(df,columnas_seleccionadas)
                     else:
                         df[['Abs 690 Triplicados 1', 'Abs 690 Triplicados 2', 'Abs 690 Triplicados 3']] = df['Abs 690 Triplicados'].str.split(expand=True)
                         df1 = tablita(df,columnas_seleccionadas)
                         tex1=self.dato2
                         self.tabla = df1.rename(columns={'Media Fosforo Total (mg/L)': tex1})
                     
                     
                 

                
                 
                 tabla1 = self.tabla
                 #tabla1 = pd.concat([tabla1, df[['Triplicados 1', 'Triplicados 2', 'Triplicados 3']],df[['Abs 690 Triplicados 1', 'Abs 690 Triplicados 2', 'Abs 690 Triplicados 3']]], axis=1)
                 print(tabla1)
                 self.tabla1 = tabla1
                 self.label_6.setText("Archivo Filtrado")

                
            except Exception as e:
                 QMessageBox.critical(self, 'Error', f'Error al filtrar la tabla: {str(e)}')
            
        else:
            QMessageBox.warning(self, 'Advertencia', 'No se seleccion� ninguna tabla.')

    def save2(self):
        if hasattr(self, 'tabla1'):
            try:
                nombre_archivo = os.path.splitext(os.path.basename(self.archivo_seleccionado))[0]
                nuevo_nombre_excel = f"Filtrado_{nombre_archivo}.xlsx"
                options = QFileDialog.Options()
                options |= QFileDialog.ReadOnly
                ruta_guardado, _ = QFileDialog.getSaveFileName(self, 'Guardar en Excel', nuevo_nombre_excel, 'Archivos Excel (*.xlsx);;Todos los archivos (*)', options=options)
            
                if ruta_guardado:
                # Crear un DataFrame a partir de self.tabla1 (aseg�rate de tener pandas instalado)
                    df = pd.DataFrame(self.tabla1)
                
                # Crear un escritor de Excel
                    writer = pd.ExcelWriter(ruta_guardado, engine='xlsxwriter')
                    df.to_excel(writer, index=False, sheet_name='Hoja1')
                
                # Obtener la hoja de c�lculo
                    workbook = writer.book
                    worksheet = workbook.get_worksheet_by_name('Hoja1')
                
                # Formatear la fila de t�tulos en verde y centrar los datos
                    header_format = workbook.add_format({'bg_color': '#00FF00', 'align': 'center', 'valign': 'vcenter','bold': True,'border':1})
                    for col_num, value in enumerate(df.columns.values):
                        worksheet.write(0, col_num, value, header_format)
                    #header_alignment = Alignment(horizontal="center", vertical="center")

                    cell_format=workbook.add_format({ 'align': 'center', 'valign': 'vcenter','border':1})
                # Ajustar autom�ticamente el ancho de las columnas
                    for i, col in enumerate(df.columns):
                        max_len = max(df[col].astype(str).str.len())
                        max_len = max(max_len, len(col))
                        worksheet.set_column(i, i, max_len + 2)
                    worksheet.set_default_row(15)  # Altura de fila predeterminada
                    worksheet.conditional_format(1, 0, len(df), len(df.columns)-1, {'type': 'no_blanks', 'format': cell_format})
                # Guardar el archivo Excel
                    writer.close()
                
                    QMessageBox.information(self, 'Informaci\u00F3n', f'Archivo Excel guardado en: {ruta_guardado}')
                    self.label_6.setText("Archivo Guardado")
                    self.bt_save2.setEnabled(False)
                    self.bt_filter2.setEnabled(False)
                else:
                    QMessageBox.warning(self, 'Advertencia', 'No se seleccion\u00F3 una ruta de guardado.')
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Error al guardar en Excel: {str(e)}')
        else:
            QMessageBox.warning(self, 'Advertencia', 'No se seleccion� ninguna tabla.')
    
          

        
if __name__=="__main__":
    app=QApplication(sys.argv)
    mi_app=rooti()
    mi_app.show()
    sys.exit(app.exec_())